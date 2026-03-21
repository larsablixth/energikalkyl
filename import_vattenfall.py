"""
Import consumption data from Vattenfall Eldistribution Excel files.

Vattenfall format (Serie sheet):
- Row 1: year
- Row 2: headers with month names in columns 12-23 (Jan-Dec)
- Row 3: column labels (col 7 = Summa/dag)
- Row 4+: one row per day
  - Col 1: month name, Col 3: day of month, Col 7: daily total kWh
  - Cols 12-23: HOURLY data (24 values per day, one column per month)
    Col 12 (L) = January, Col 13 (M) = February, ..., Col 23 (W) = December
    Each column has 24 consecutive values per day (hour 00-23)
"""

import openpyxl
from datetime import date
from pathlib import Path


MONTH_MAP = {
    "jan": 1, "jan.": 1, "januari": 1,
    "feb": 2, "feb.": 2, "februari": 2,
    "mar": 3, "mar.": 3, "mars": 3,
    "apr": 4, "apr.": 4, "april": 4,
    "maj": 5, "maj.": 5,
    "jun": 6, "jun.": 6, "juni": 6,
    "jul": 7, "jul.": 7, "juli": 7,
    "aug": 8, "aug.": 8, "augusti": 8,
    "sep": 9, "sep.": 9, "sept": 9, "sept.": 9, "september": 9,
    "okt": 10, "okt.": 10, "oktober": 10,
    "nov": 11, "nov.": 11, "november": 11,
    "dec": 12, "dec.": 12, "december": 12,
}


def parse_vattenfall_excel(path: str) -> list[dict]:
    """
    Parse a Vattenfall Eldistribution Excel file.
    Returns list of dicts with date, consumption_kwh (daily totals).
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    if "Serie" not in wb.sheetnames:
        raise ValueError(f"Hittade inte 'Serie'-bladet i {path}. Blad: {wb.sheetnames}")

    ws = wb["Serie"]
    rows = list(ws.iter_rows(values_only=True))

    # Row 0: year
    year = None
    for cell in rows[0]:
        if isinstance(cell, (int, float)) and 2000 <= cell <= 2100:
            year = int(cell)
            break
    if year is None:
        raise ValueError("Kunde inte hitta årstal i filen")

    # Parse data rows (row 3+)
    # Column 0 = month name, column 2 = day of month, column 6 = daily total kWh
    results = []
    for row_idx in range(3, len(rows)):
        row = rows[row_idx]
        if row is None:
            continue

        month_cell = row[0]
        day_cell = row[2]

        if not isinstance(month_cell, str):
            continue
        month_str = month_cell.strip().lower()
        if month_str not in MONTH_MAP:
            continue
        month_num = MONTH_MAP[month_str]

        if not isinstance(day_cell, (int, float)) or day_cell < 1 or day_cell > 31:
            continue
        day = int(day_cell)

        # Use column 6 (Summa/dag) for daily total
        value = row[6] if len(row) > 6 else None
        if value is None or not isinstance(value, (int, float)):
            continue

        try:
            d = date(year, month_num, day)
        except ValueError:
            continue

        results.append({
            "date": d.isoformat(),
            "consumption_kwh": round(float(value), 4),
        })

    wb.close()
    print(f"  År: {year}, {len(results)} dagar med data")
    return results


def parse_vattenfall_hourly(path: str) -> list[dict]:
    """
    Extract hourly consumption from Vattenfall Excel (Serie sheet columns L-W).

    Returns list of dicts with date, hour (0-23), kwh.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Serie" not in wb.sheetnames:
        wb.close()
        return []

    ws = wb["Serie"]

    # Get year from row 1
    year = None
    for cell in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for v in cell:
            if isinstance(v, (int, float)) and 2000 <= v <= 2100:
                year = int(v)
                break

    if year is None:
        wb.close()
        return []

    # Columns 12-23 (L-W) = Jan-Dec hourly data
    # Each column: row 3 = first value (header/hour 0 of day 1)
    # Then 24 values per day
    import calendar
    hourly = []
    for month_offset, col in enumerate(range(12, 24)):
        month = month_offset + 1
        days_in_month = calendar.monthrange(year, month)[1]

        # Read all values in this column
        vals = []
        for r in range(3, ws.max_row + 1):
            v = ws.cell(r, col).value
            if v is not None and isinstance(v, (int, float)):
                vals.append(float(v))

        if not vals:
            continue

        # First value might be a monthly average — skip if count matches days*24 + 1
        if len(vals) == days_in_month * 24 + 1:
            vals = vals[1:]
        elif len(vals) < days_in_month * 24:
            # Partial month or different format — take what we have
            pass

        # Extract 24 values per day
        for day_idx in range(min(days_in_month, len(vals) // 24)):
            try:
                d = date(year, month, day_idx + 1)
            except ValueError:
                continue
            for hour in range(24):
                idx = day_idx * 24 + hour
                if idx < len(vals):
                    hourly.append({
                        "date": d.isoformat(),
                        "hour": hour,
                        "kwh": round(vals[idx], 3),
                    })

    wb.close()
    print(f"  År: {year}, {len(hourly)} timvärden ({len(hourly)//24} dagar)")
    return hourly


def load_vattenfall_files(*paths: str) -> list[dict]:
    """Load and merge multiple Vattenfall Excel files."""
    all_data = []
    for path in paths:
        print(f"  Läser {Path(path).name}...")
        data = parse_vattenfall_excel(path)
        all_data.extend(data)

    # Sort and deduplicate
    all_data.sort(key=lambda r: r["date"])
    seen = set()
    unique = []
    for r in all_data:
        if r["date"] not in seen:
            seen.add(r["date"])
            unique.append(r)

    if unique:
        total_kwh = sum(r["consumption_kwh"] for r in unique)
        num_days = len(unique)
        print(f"  Totalt: {num_days} dagar, {unique[0]['date']} → {unique[-1]['date']}")
        print(f"  Total förbrukning: {total_kwh:,.0f} kWh")
        print(f"  Snitt: {total_kwh/num_days:.0f} kWh/dag | {total_kwh/num_days*365.25:,.0f} kWh/år")

    return unique


def load_vattenfall_hourly(*paths: str) -> list[dict]:
    """Load and merge hourly data from multiple Vattenfall Excel files."""
    all_hourly = []
    for path in paths:
        print(f"  Läser timdata från {Path(path).name}...")
        all_hourly.extend(parse_vattenfall_hourly(path))

    # Sort and deduplicate by (date, hour)
    all_hourly.sort(key=lambda r: (r["date"], r["hour"]))
    seen = set()
    unique = []
    for r in all_hourly:
        key = (r["date"], r["hour"])
        if key not in seen:
            seen.add(key)
            unique.append(r)

    if unique:
        total = sum(r["kwh"] for r in unique)
        days = len(set(r["date"] for r in unique))
        print(f"  Totalt: {len(unique)} timvärden, {days} dagar, {total:,.0f} kWh")

    return unique


def vattenfall_hourly_to_seasonal_profile(hourly_data: list[dict]) -> dict[int, dict[int, float]]:
    """Build seasonal hourly profile directly from Vattenfall hourly data.

    Returns dict: month -> hour -> average kW
    """
    from collections import defaultdict
    sums = defaultdict(lambda: defaultdict(float))
    counts = defaultdict(lambda: defaultdict(int))

    for r in hourly_data:
        m = int(r["date"].split("-")[1])
        h = r["hour"]
        sums[m][h] += r["kwh"]  # kWh per hour ≈ kW average
        counts[m][h] += 1

    profile = {}
    for m in range(1, 13):
        profile[m] = {}
        for h in range(24):
            if counts[m][h] > 0:
                profile[m][h] = round(sums[m][h] / counts[m][h], 3)
            else:
                profile[m][h] = 0.0

    return profile


def vattenfall_to_monthly_profile(data: list[dict]) -> dict[int, float]:
    """
    Build average daily consumption per month from Vattenfall daily data.
    Returns dict mapping month (1-12) to average kWh/day.
    """
    from collections import defaultdict
    monthly_sums = defaultdict(float)
    monthly_counts = defaultdict(int)

    for r in data:
        m = int(r["date"].split("-")[1])
        monthly_sums[m] += r["consumption_kwh"]
        monthly_counts[m] += 1

    profile = {}
    for m in range(1, 13):
        if monthly_counts[m] > 0:
            profile[m] = monthly_sums[m] / monthly_counts[m]
        else:
            profile[m] = 0.0
    return profile


def vattenfall_to_seasonal_profile(daily_data: list[dict], hourly_shape: dict[int, float] = None) -> dict[int, dict[int, float]]:
    """
    Build seasonal hourly profile from Vattenfall daily data.

    Uses monthly averages from Vattenfall to scale an hourly shape.
    If no hourly_shape is provided, uses a default residential profile.

    Returns dict: month -> hour -> kW
    """
    monthly_daily = vattenfall_to_monthly_profile(daily_data)

    # Default hourly shape if none provided (typical Swedish household)
    if hourly_shape is None:
        hourly_shape = {
            0: 3.0, 1: 2.5, 2: 2.0, 3: 2.0, 4: 2.0, 5: 2.5,
            6: 3.5, 7: 4.0, 8: 3.5, 9: 3.0, 10: 2.5, 11: 2.5,
            12: 2.5, 13: 2.5, 14: 2.5, 15: 3.0, 16: 3.5, 17: 4.0,
            18: 4.5, 19: 4.0, 20: 3.5, 21: 3.0, 22: 3.0, 23: 3.0,
        }

    base_daily_kwh = sum(hourly_shape.values())
    if base_daily_kwh == 0:
        return {m: {h: 0.0 for h in range(24)} for m in range(1, 13)}

    seasonal = {}
    for m in range(1, 13):
        if monthly_daily[m] > 0:
            scale = monthly_daily[m] / base_daily_kwh
        else:
            scale = 1.0
        seasonal[m] = {h: hourly_shape[h] * scale for h in range(24)}

    return seasonal


def print_vattenfall_summary(data: list[dict]):
    """Print monthly consumption summary."""
    months_sv = ["", "Jan", "Feb", "Mar", "Apr", "Maj", "Jun",
                 "Jul", "Aug", "Sep", "Okt", "Nov", "Dec"]
    profile = vattenfall_to_monthly_profile(data)

    print(f"\n  Vattenfall förbrukningsprofil:")
    print(f"  {'Månad':>6} {'kWh/dag':>10} {'kWh/mån':>10}")
    print(f"  {'-'*30}")
    for m in range(1, 13):
        kwh = profile[m]
        kwh_month = kwh * 30.44
        bar = "█" * int(kwh / 3)
        print(f"  {months_sv[m]:>5}  {kwh:>8.1f}  {kwh_month:>8.0f}  {bar}")
    total = sum(profile[m] * 30.44 for m in range(1, 13))
    print(f"  {'-'*30}")
    print(f"  {'Total':>5}  {total/365.25:>8.1f}  {total:>8.0f}")
